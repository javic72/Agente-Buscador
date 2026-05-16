from datetime import datetime
from pathlib import Path
from urllib.parse import urlsplit, urlunsplit

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADERS = [
    "Fuente",
    "numero de apariciones",
    "sectores",
    "ultima aparicion",
    "prioridad maxima encontrada",
]

PRIORITY_RANK = {
    "Descartada": 0,
    "Baja": 1,
    "Media": 2,
    "Alta": 3,
}


def update_source_catalog(xlsx_path, opportunities):
    xlsx_path = Path(xlsx_path)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    try:
        workbook, sheet = load_or_create_workbook(xlsx_path)
    except PermissionError:
        print(
            f"  Aviso: el catálogo de fuentes está bloqueado y no se pudo actualizar: {xlsx_path}"
        )
        return xlsx_path

    existing_rows = read_existing_rows(sheet)

    for opportunity in opportunities:
        main_source = get_main_source_url(opportunity)
        if not main_source:
            continue

        row = existing_rows.setdefault(
            main_source,
            {
                "Fuente": main_source,
                "numero de apariciones": 0,
                "sectores": "",
                "ultima aparicion": "",
                "prioridad maxima encontrada": "Descartada",
            },
        )

        row["numero de apariciones"] = int(row.get("numero de apariciones") or 0) + 1
        row["sectores"] = merge_sector(row.get("sectores", ""), opportunity.get("sector", ""))
        row["ultima aparicion"] = latest_date(
            row.get("ultima aparicion", ""),
            opportunity.get("detected_date", "") or datetime.now().date().isoformat(),
        )
        row["prioridad maxima encontrada"] = max_priority(
            row.get("prioridad maxima encontrada", "Descartada"),
            opportunity.get("priority", "Descartada"),
        )

    write_rows(sheet, existing_rows)
    format_sheet(sheet)
    try:
        workbook.save(xlsx_path)
    except PermissionError:
        print(
            f"  Aviso: el catálogo de fuentes está bloqueado y no se pudo guardar: {xlsx_path}"
        )
    return xlsx_path


def load_or_create_workbook(xlsx_path):
    if xlsx_path.exists():
        workbook = load_workbook(xlsx_path)
        sheet = workbook.active
        sheet.title = "Fuentes"
        ensure_headers(sheet)
        return workbook, sheet

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Fuentes"
    sheet.append(HEADERS)
    return workbook, sheet


def ensure_headers(sheet):
    current_headers = [sheet.cell(row=1, column=index).value for index in range(1, len(HEADERS) + 1)]
    if current_headers != HEADERS:
        for index, header in enumerate(HEADERS, start=1):
            sheet.cell(row=1, column=index).value = header


def read_existing_rows(sheet):
    rows = {}
    for values in sheet.iter_rows(min_row=2, values_only=True):
        if not values or not values[0]:
            continue
        row = dict(zip(HEADERS, values))
        rows[row["Fuente"]] = row
    return rows


def write_rows(sheet, rows):
    if sheet.max_row > 1:
        sheet.delete_rows(2, sheet.max_row - 1)

    sorted_rows = sorted(
        rows.values(),
        key=lambda row: (
            -int(row.get("numero de apariciones") or 0),
            str(row.get("Fuente", "")),
        ),
    )
    for row in sorted_rows:
        sheet.append([row.get(header, "") for header in HEADERS])


def format_sheet(sheet):
    header_fill = PatternFill("solid", fgColor="17202A")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    widths = {
        "A": 42,
        "B": 22,
        "C": 46,
        "D": 18,
        "E": 28,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    if sheet.max_row >= 2:
        table_ref = f"A1:E{sheet.max_row}"
        sheet.tables.clear()
        table = Table(displayName="FuentesDetectadas", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)


def get_main_source_url(opportunity):
    candidates = [
        opportunity.get("source_url", ""),
        opportunity.get("url", ""),
    ]
    for candidate in candidates:
        main_url = normalize_main_url(candidate)
        if main_url and "news.google.com" not in main_url:
            return main_url
    return ""


def normalize_main_url(url):
    if not url:
        return ""

    parts = urlsplit(url.strip())
    if not parts.netloc:
        return ""

    scheme = parts.scheme or "https"
    netloc = parts.netloc.lower()
    return urlunsplit((scheme, netloc, "/", "", ""))


def merge_sector(existing, new_sector):
    sectors = [sector.strip() for sector in str(existing or "").split(";") if sector.strip()]
    if new_sector and new_sector not in sectors:
        sectors.append(new_sector)
    return "; ".join(sorted(sectors))


def latest_date(existing, candidate):
    if not existing:
        return candidate
    if not candidate:
        return existing
    return max(str(existing), str(candidate))


def max_priority(existing, candidate):
    existing = existing or "Descartada"
    candidate = candidate or "Descartada"
    if PRIORITY_RANK.get(candidate, 0) > PRIORITY_RANK.get(existing, 0):
        return candidate
    return existing
